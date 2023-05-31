import json
from openpyxl import load_workbook


# Открываем файл
workbook = load_workbook('example.xlsx')
# Выбираем нужный лист
sheet = workbook.active
# Создаем пустой список для данных
data = []

# Проходимся по всем строкам и столбцам и добавляем данные в список
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Преобразуем данные в словарь
keys = data[0]
values = data[1:]
dict_data = [dict(zip(keys, row)) for row in values]

# Преобразуем словарь в json и сохраняем в файл
with open('example.json', 'w') as f:
    json.dump(dict_data, f)
