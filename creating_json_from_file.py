import json
import base64
from openpyxl import load_workbook


# Открываем файл и открываем нужный лист
workbook = load_workbook('areas_import_template_4проверки-dairy.xlsx')
sheet = workbook.active

# Создаем пустой список для данных, проходимся по всем строкам и добавляем данные в список
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Преобразуем данные в словарь
keys = data[0]
values = data[1:]
dict_data = [dict(zip(keys, row)) for row in values]

# Преобразуем словарь в json и сохраняем в файл
with open('file.json', 'w') as f:
    json.dump(dict_data, f)
